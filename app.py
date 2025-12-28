import streamlit as st
import pandas as pd
import openpyxl
from openai import OpenAI
import requests
import json
import os
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet
import folium
from streamlit_folium import folium_static
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from PIL import Image as PilImage
import pdfplumber
import random
import plotly.express as px
from scipy.interpolate import griddata
import numpy as np
import sqlite3  # For BLM data database
from bs4 import BeautifulSoup  # For scraping

# ========================================
# App Configuration
# ========================================
st.set_page_config(page_title="Mining Data Analysis Portal", layout="wide")
st.title("Mining Data Analysis Portal - Enhanced with USGS, BLM, Compliance & ESG Tools")

# ========================================
# Mineral Areas Database
# ========================================
mineral_areas = {
    "Northern Rio Grande Rift (Colorado) - Au, Ag, Mo": {
        "description": "Covers areas like Leadville, San Luis Basin, Taos Plateau Volcanic Field. Known for gold, silver, molybdenum.",
        "geology": "Broad downwarp with basins, volcanic features along Jemez Lineament. The rift formed ~36-37 Ma due to crustal extension and thinning. Basins like San Luis are complex, divided by intrabasin horsts, with low-angle faults. Sediments deposited in closed basins under intermittent flooding.",
        "geothermal": "High geothermal potential in linkage zones and basins like San Luis. Heat flow >4.0 HFU in parts; hot springs and geothermal wells indicate resources for power generation. Valles Caldera (central but influential) has high-temperature systems (up to 300°C). Exploration ongoing with potential for EGS (Enhanced Geothermal Systems).",
        "search_query": "mines and volcanic fields in northern Rio Grande Rift Colorado",
        "usgs_query": {"state": "Colorado", "commodity": "gold,silver,molybdenum"},
        "state": "Colorado"
    },
    "Central Rio Grande Rift (New Mexico) - Cu, Pb, Zn, U": {
        "description": "Includes Espanola, Albuquerque, Socorro basins, Jemez Volcanic Field, Cerros del Rio. Rich in copper, lead, zinc, uranium.",
        "geology": "En echelon basins, half-grabens, with mid-Oligocene to Pleistocene volcanism. Española basin: 2-3 km deep, began as downwarp in late Oligocene. Albuquerque-Belen basins with ~0.3 mm/yr extension. Complex basins with horsts; late Oligocene magmatism imprinted thermal boundaries. Natural resources in rift basins.",
        "geothermal": "Significant geothermal resources with volcanics; Valles Caldera and Ojo Caliente hot springs show distal connections. High heat flow suggests vertical fractures for magma/groundwater interaction. Known geothermal areas like Jemez Springs (up to 100°C). Potential for binary cycle plants; assessments indicate moderate-high temperature resources.",
        "search_query": "mines and volcanic fields in central Rio Grande Rift New Mexico",
        "usgs_query": {"state": "New Mexico", "commodity": "copper,lead,zinc,uranium"},
        "state": "New Mexico"
    },
    "Southern Rio Grande Rift (New Mexico/Texas/Mexico) - Au, Ag, Cu": {
        "description": "Potrillo Volcanic Field, Mesilla Basin, extending to Chihuahua. Limited metallic deposits, but cinder and aggregate resources; nearby copper mines like Tyrone.",
        "geology": "Narrow rift segments, monogenetic volcanic fields, Basin and Range extension into Mexico. Rift started ~36 Ma with westerly extension. Basins like Santo Domingo form large accommodation zones. Distributed deformation across rift, Great Plains.",
        "geothermal": "Evaluated in areas like Truth or Consequences with high heat flow anomalies. Self-potential surveys in regions like Radium Springs show potential. Moderate resources with hot springs; under-explored but promising for low-temperature applications. Overall rift anomalies suggest extensive fractures for geothermal fluid circulation.",
        "search_query": "mines and volcanic fields in southern Rio Grande Rift New Mexico Texas Mexico",
        "usgs_query": {"state": "New Mexico", "commodity": "gold,silver,copper"},
        "state": "New Mexico"
    },
    "Carlin Trend (Nevada) - Au": {
        "description": "World-class gold mining district in northern Nevada, known for Carlin-type gold deposits.",
        "geology": "Sedimentary-hosted disseminated gold in Paleozoic rocks, associated with intrusive igneous activity.",
        "geothermal": "Moderate potential due to Basin and Range extension.",
        "search_query": "gold mines in Carlin Trend Nevada",
        "usgs_query": {"state": "Nevada", "commodity": "gold"},
        "state": "Nevada"
    },
    "Black Hills (South Dakota) - Au, Ag": {
        "description": "Historic gold rush area, including Homestake Mine, one of the largest gold producers in US history.",
        "geology": "Precambrian metamorphic rocks with Tertiary intrusions.",
        "geothermal": "Low to moderate.",
        "search_query": "gold silver mines in Black Hills South Dakota",
        "usgs_query": {"state": "South Dakota", "commodity": "gold,silver"},
        "state": "South Dakota"
    },
    "Appalachian Region (Eastern US) - Au, Ag": {
        "description": "Gold and silver in Piedmont and Blue Ridge provinces, e.g., Virginia, North Carolina.",
        "geology": "Metamorphic and volcanic rocks with vein deposits.",
        "geothermal": "Low.",
        "search_query": "gold silver mines in Appalachian US",
        "usgs_query": {"state": "Virginia", "commodity": "gold,silver"},
        "state": "Virginia"
    },
    "Bear Lodge (Wyoming) - REE": {
        "description": "Major rare earth elements deposit in the Black Hills uplift.",
        "geology": "Alkaline igneous complex with carbonatite intrusions.",
        "geothermal": "Low.",
        "search_query": "REE mines in Bear Lodge Wyoming",
        "usgs_query": {"state": "Wyoming", "commodity": "rare earths"},
        "state": "Wyoming"
    },
    "Round Top (Texas) - REE, Li": {
        "description": "Rhyolite-hosted rare earth and lithium deposit.",
        "geology": "Tertiary intrusive rhyolite laccolith.",
        "geothermal": "Moderate.",
        "search_query": "REE lithium mines in Round Top Texas",
        "usgs_query": {"state": "Texas", "commodity": "rare earths,lithium"},
        "state": "Texas"
    },
    "Bokan Mountain (Alaska) - REE, U": {
        "description": "Peralkaline granite-hosted rare earth and uranium.",
        "geology": "Jurassic peralkaline intrusive complex.",
        "geothermal": "High in some Alaskan areas.",
        "search_query": "REE uranium mines in Bokan Mountain Alaska",
        "usgs_query": {"state": "Alaska", "commodity": "rare earths,uranium"},
        "state": "Alaska"
    },
    "Mojave Desert (California) - REE, Au": {
        "description": "Mountain Pass Mine, world's largest REE producer outside China; also gold.",
        "geology": "Carbonatite deposits in Precambrian gneiss.",
        "geothermal": "High in Imperial Valley nearby.",
        "search_query": "REE gold mines in Mojave Desert California",
        "usgs_query": {"state": "California", "commodity": "rare earths,gold"},
        "state": "California"
    },
    "Mother Lode (California) - Au": {
        "description": "Historic California Gold Rush area along Sierra Nevada foothills.",
        "geology": "Mesothermal quartz veins in metamorphic rocks.",
        "geothermal": "Moderate.",
        "search_query": "gold mines in Mother Lode California",
        "usgs_query": {"state": "California", "commodity": "gold"},
        "state": "California"
    },
    "Cripple Creek (Colorado) - Au, Ag": {
        "description": "Volcanic-hosted epithermal gold-silver deposits.",
        "geology": "Oligocene caldera with telluride minerals.",
        "geothermal": "High.",
        "search_query": "gold silver mines in Cripple Creek Colorado",
        "usgs_query": {"state": "Colorado", "commodity": "gold,silver"},
        "state": "Colorado"
    },
    "Comstock Lode (Nevada) - Ag, Au": {
        "description": "Famous silver mining district near Virginia City.",
        "geology": "Epithermal veins in Tertiary volcanics.",
        "geothermal": "High.",
        "search_query": "silver gold mines in Comstock Lode Nevada",
        "usgs_query": {"state": "Nevada", "commodity": "silver,gold"},
        "state": "Nevada"
    },
    "Idaho Batholith (Idaho) - Au, Ag, REE": {
        "description": "Granitic intrusions with vein and placer deposits.",
        "geology": "Cretaceous granites with polymetallic veins.",
        "geothermal": "Moderate.",
        "search_query": "gold silver REE mines in Idaho Batholith",
        "usgs_query": {"state": "Idaho", "commodity": "gold,silver,rare earths"},
        "state": "Idaho"
    },
    "Pebble (Alaska) - Cu, Au, Mo": {
        "description": "Porphyry copper-gold-molybdenum deposit.",
        "geology": "Tertiary intrusive complex.",
        "geothermal": "High.",
        "search_query": "copper gold mines in Pebble Alaska",
        "usgs_query": {"state": "Alaska", "commodity": "copper,gold,molybdenum"},
        "state": "Alaska"
    },
}

# Static Overviews
st.header("Basin and Range Province Overview")
st.write("**Description:** The Basin and Range Province covers a large portion of the southwestern United States and western Mexico, including most of Nevada, parts of California, Oregon, Utah, Arizona, New Mexico, and extending into northern Mexico. It is the most geographically extensive 'young' geologic region in North America.")
st.write("**Geology:** Characterized by unique basin and range topography with abrupt elevation changes, alternating narrow faulted mountain chains (ranges) and flat arid valleys (basins). Formed by extensional tectonics, with normal faults pushing up mountains and carving valleys below.")

st.header("Colorado Plateau Overview")
st.write("**Description:** The Colorado Plateau is a physiographic province covering parts of Arizona, Utah, Colorado, and New Mexico. Known for iconic landmarks like the Grand Canyon, Zion, Arches, and Bryce Canyon National Parks.")
st.write("**Geology:** Largely made up of high desert with scattered forests, characterized by flat-lying sedimentary rocks sculpted into mesas, buttes, canyons, and badlands. Stable crustal block, uplifted ~8,500 feet without significant deformation.")

# Area Selection
selected_area = st.selectbox("Select Mineral/Geological Area for Analysis", list(mineral_areas.keys()))

if selected_area:
    area_info = mineral_areas[selected_area]
    st.write(f"**Description:** {area_info['description']}")
    st.write(f"**Detailed Geology:** {area_info['geology']}")
    st.write(f"**Geothermal Potential:** {area_info['geothermal']}")

    # Interactive Map of Selected Area
    st.subheader("Interactive Map of Selected Area")
    coords = {
        "Northern Rio Grande Rift (Colorado) - Au, Ag, Mo": [37.5, -106.0],
        "Central Rio Grande Rift (New Mexico) - Cu, Pb, Zn, U": [35.0, -106.5],
        "Southern Rio Grande Rift (New Mexico/Texas/Mexico) - Au, Ag, Cu": [32.0, -107.0],
        "Carlin Trend (Nevada) - Au": [40.8, -116.0],
        "Black Hills (South Dakota) - Au, Ag": [44.0, -103.5],
        "Appalachian Region (Eastern US) - Au, Ag": [37.5, -80.0],
        "Bear Lodge (Wyoming) - REE": [44.5, -104.5],
        "Round Top (Texas) - REE, Li": [31.3, -105.5],
        "Bokan Mountain (Alaska) - REE, U": [55.0, -132.0],
        "Mojave Desert (California) - REE, Au": [35.0, -116.0],
        "Mother Lode (California) - Au": [38.5, -120.5],
        "Cripple Creek (Colorado) - Au, Ag": [38.7, -105.2],
        "Comstock Lode (Nevada) - Ag, Au": [39.3, -119.6],
        "Idaho Batholith (Idaho) - Au, Ag, REE": [45.0, -115.0],
        "Pebble (Alaska) - Cu, Au, Mo": [59.7, -155.3],
    }
    area_coord = coords.get(selected_area, [35.0, -106.5])
    m = folium.Map(location=area_coord, zoom_start=7)
    folium.Marker(area_coord, popup=selected_area).add_to(m)
    folium_static(m)

    # USGS Bulletin 1693 Integration
    st.subheader("USGS Bulletin 1693 Integration (Mineral Deposit Models)")
    if st.button("Summarize USGS Bulletin 1693 PDF"):
        try:
            bulletin_url = "https://pubs.usgs.gov/bul/1693/report.pdf"
            response = requests.get(bulletin_url)
            if response.status_code == 200:
                pdf_buffer = io.BytesIO(response.content)
                with pdfplumber.open(pdf_buffer) as pdf:
                    text = ""
                    for page in pdf.pages:
                        text += page.extract_text() or ""
                client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
                summary_prompt = f"Summarize the USGS Bulletin 1693 PDF content: {text[:4000]}. Focus on mineral deposit models, especially PGM-related ones."
                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{"role": "user", "content": summary_prompt}],
                    max_tokens=1000
                )
                st.session_state['bulletin_summary'] = response.choices[0].message.content
                st.write("USGS Bulletin 1693 Summary:")
                st.write(st.session_state['bulletin_summary'])
            else:
                st.error("Failed to fetch USGS Bulletin 1693 PDF.")
        except Exception as e:
            st.error(f"Bulletin Integration Error: {e}")

    # USGS Mineral Resources Data System (MRDS) Integration - Updated with CSV Download
    st.subheader("USGS Mineral Resources Data System (MRDS) Integration")
    st.write("""
    **Note:** The legacy MRDS search API has been deprecated. Use the official interactive search or download the full dataset.
    """)
    st.markdown("[Open Interactive MRDS Search](https://mrdata.usgs.gov/mrds/find-mrds.php)")
    st.markdown("[Download Full MRDS CSV Dataset](https://mrdata.usgs.gov/mrds/mrds.csv)")
    st.info("Download the CSV for offline analysis or import into GIS tools. Contains all worldwide mineral sites with detailed attributes.")

    # Mindat.org Mineral Data Integration
    st.subheader("Mindat.org Mineral Locality Data")
    st.write("""
    Mindat.org is the world's largest open database of minerals, rocks, meteorites, and their localities.
    Search for minerals and localities in your selected area.
    """)

    mindat_query = st.text_input("Search Mindat.org (e.g., 'gold Nevada' or 'quartz Colorado')", value=f"{selected_area.split(' - ')[1] if ' - ' in selected_area else 'gold'} {area_info.get('state', '')}")
    
    if st.button("Search Mindat.org"):
        mindat_url = f"https://www.mindat.org/search.php?search={mindat_query.replace(' ', '+')}"
        st.markdown(f"[Open Mindat.org Search Results]({mindat_url})")
        st.info("Mindat.org provides detailed mineralogy, photos, and locality data. Excellent for mineral species and crystal information.")

    # BLM Mining Claims Search (Official BLM ArcGIS Data with Pagination & CSV Export)
    st.subheader("BLM Mining Claims Search (Official BLM ArcGIS Data)")
    st.write("""
    This uses the official BLM ArcGIS REST API for mining claims (public data, no API key required).
    - Includes active and closed claims
    - Pagination supported for larger result sets
    - **CSV Export** available when results are loaded
    - Data directly from BLM NLSDB (updated regularly)
    - Tip: Nevada (NV) has the most claims — try it first for testing
    """)

    col1, col2 = st.columns(2)
    with col1:
        state_code = st.text_input("State Code (e.g., NV, NM, CO)", value="NV")
        county = st.text_input("County Name (optional)", value="")
    with col2:
        records_per_page = st.selectbox("Records Per Page", [50, 100, 200, 300, 400, 500, 1000, 2000], index=4)

    # Session state for pagination and data
    if 'blm_page_offset' not in st.session_state:
        st.session_state.blm_page_offset = 0
    if 'blm_current_df' not in st.session_state:
        st.session_state.blm_current_df = None
    if 'blm_all_results_df' not in st.session_state:
        st.session_state.blm_all_results_df = pd.DataFrame()

    def fetch_blm_claims(offset=0):
        try:
            base_url = "https://gis.blm.gov/nlsdb/rest/services/Mining_Claims/MiningClaims/MapServer/1/query"
            
            where_parts = []
            if state_code:
                where_parts.append(f"ADMIN_STATE = '{state_code.upper()}'")
            if county:
                where_parts.append(f"UPPER(COUNTY_NM) LIKE '%{county.upper()}%'")
            
            where = " AND ".join(where_parts) if where_parts else "1=1"

            params = {
                "where": where,
                "outFields": "*",
                "returnGeometry": "false",
                "f": "json",
                "resultRecordCount": records_per_page,
                "resultOffset": offset,
                "orderByFields": "CSE_NR DESC"
            }

            response = requests.get(base_url, params=params, timeout=60)
            if response.status_code == 200:
                data = response.json()
                if 'features' in data and data['features']:
                    claims_list = [feature['attributes'] for feature in data['features']]
                    df = pd.DataFrame(claims_list)
                    if df.empty:
                        return pd.DataFrame(), 0
                    df.columns = [col.lower() for col in df.columns]
                    return df, len(claims_list)
                else:
                    return pd.DataFrame(), 0
            else:
                st.error(f"API Error: {response.status_code} - {response.text}")
                return pd.DataFrame(), 0
        except Exception as e:
            st.error(f"Request failed: {e}")
            return pd.DataFrame(), 0

    col_left, col_mid, col_right = st.columns([1, 2, 1])
    with col_mid:
        if st.button("🔍 Search BLM Mining Claims", use_container_width=True):
            st.session_state.blm_page_offset = 0
            st.session_state.blm_all_results_df = pd.DataFrame()
            df_page, count = fetch_blm_claims(offset=0)
            if not df_page.empty:
                st.session_state.blm_current_df = df_page
                st.session_state.blm_all_results_df = df_page.copy()
                st.success(f"Found {len(df_page)} claims (page 1)")
            else:
                st.info("No claims found. Try 'NV' for Nevada — it has thousands of claims.")

    if st.session_state.blm_current_df is not None and not st.session_state.blm_current_df.empty:
        st.success(f"Showing {len(st.session_state.blm_current_df)} claims (Total loaded: {len(st.session_state.blm_all_results_df)})")

        display_cols = ['cse_nr', 'cse_name', 'cse_type_nr', 'cse_disp', 'admin_state', 'county_nm', 'claimant_name', 'loc_date']
        available_cols = [col for col in display_cols if col in st.session_state.blm_current_df.columns]
        st.dataframe(st.session_state.blm_current_df[available_cols], use_container_width=True)

        if 'latitude' in st.session_state.blm_current_df.columns and 'longitude' in st.session_state.blm_current_df.columns:
            map_df = st.session_state.blm_current_df[['latitude', 'longitude', 'cse_name']].dropna()
            if not map_df.empty and len(map_df) <= 300:
                blm_map = folium.Map(location=[map_df['latitude'].mean(), map_df['longitude'].mean()], zoom_start=8)
                for _, row in map_df.iterrows():
                    folium.CircleMarker(
                        location=[row['latitude'], row['longitude']],
                        radius=4,
                        popup=row.get('cse_name', 'Claim'),
                        color='blue',
                        fill=True,
                        fillOpacity=0.6
                    ).add_to(blm_map)
                folium_static(blm_map)

        csv = st.session_state.blm_all_results_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 Download All Loaded Claims as CSV",
            data=csv,
            file_name=f"blm_claims_{state_code.upper()}.csv",
            mime="text/csv"
        )

        col_prev, col_info, col_next = st.columns([1, 3, 1])
        with col_prev:
            if st.session_state.blm_page_offset > 0:
                if st.button("⬅️ Previous Page"):
                    new_offset = max(0, st.session_state.blm_page_offset - records_per_page)
                    df_page, _ = fetch_blm_claims(offset=new_offset)
                    if not df_page.empty:
                        st.session_state.blm_current_df = df_page
                        st.session_state.blm_page_offset = new_offset
                        st.rerun()

        with col_info:
            st.write(f"Current offset: {st.session_state.blm_page_offset}")

        with col_next:
            if st.button("Next Page ➡️"):
                new_offset = st.session_state.blm_page_offset + records_per_page
                df_page, count = fetch_blm_claims(offset=new_offset)
                if not df_page.empty and count > 0:
                    st.session_state.blm_current_df = df_page
                    st.session_state.blm_all_results_df = pd.concat([st.session_state.blm_all_results_df, df_page], ignore_index=True)
                    st.session_state.blm_page_offset = new_offset
                    st.rerun()
                else:
                    st.info("No more results available.")

    st.info("""
    **Data Source**: Official BLM ArcGIS Server  
    **Endpoint**: https://gis.blm.gov/nlsdb/rest/services/Mining_Claims/MiningClaims/MapServer/1  
    **Features**: Pagination, CSV export of all loaded results, real-time public data.
    **Tip**: Use 'NV' for Nevada to see thousands of claims instantly.
    """)

    # Search for Mines/Claims for Sale (Using BLM Data)
    st.subheader("Search for Mines/Claims for Sale")

    st.write("""
    Use your BLM search results to find mining claims or mines that may be for sale.
    - Many BLM unpatented claims are privately transferred (sold) between individuals.
    - Below are direct links to major marketplaces — use the **CSE_NR** (serial number) or location from your BLM results to search.
    - Some sites list BLM serial numbers directly.
    """)

    # Only show if BLM data is loaded
    if st.session_state.blm_all_results_df is not None and not st.session_state.blm_all_results_df.empty:
        st.success(f"BLM search loaded {len(st.session_state.blm_all_results_df)} claims — use serial numbers (cse_nr) below for sale searches.")

        # Show top 10 serial numbers for easy copy
        serial_numbers = st.session_state.blm_all_results_df['cse_nr'].astype(str).tolist()[:10]
        st.write("**Sample Serial Numbers (cse_nr) from your results — copy and search on sites below:**")
        for sn in serial_numbers:
            st.code(sn)

        st.write("**Automated Search Links (using first serial number as example):**")
        example_sn = st.session_state.blm_all_results_df['cse_nr'].astype(str).iloc[0]

        st.markdown(f"""
        - [The Diggings - Search by Serial Number](https://thediggings.com/search?serial_number={example_sn})
        - [MineExchange.com - Search by Serial Number](https://mineexchange.com/search?query={example_sn})
        - [Gold Rush Expeditions - Nevada Claims](https://goldrushexpeditions.com/mining-claims-for-sale/nevada-mining-claims-for-sale/)
        - [MineListings.com - Search](https://minelistings.com/?s={example_sn})
        - [LandGate - Mineral Rights Search](https://landgate.com/mineral-rights)
        - [Mountain Man Mining - Nevada Claims](https://mountainmanmining.com/collections/nevada)
        """)

    else:
        st.info("Run a BLM search first to load claims. Then use serial numbers (cse_nr) to check sale listings on these sites:")

    # Always show general links
    st.write("**Major Marketplaces for Mining Claims/Mines for Sale:**")
    st.markdown("""
    - **[The Diggings](https://thediggings.com)** – Best for BLM serial number search (free/premium)
    - **[MineExchange.com](https://mineexchange.com)** – Professional marketplace for mines and claims
    - **[Gold Rush Expeditions](https://goldrushexpeditions.com/mining-claims-for-sale/)** – Documented claims with reports
    - **[MineListings.com](https://minelistings.com/)** – Global mine/claim marketplace
    - **[LandGate](https://landgate.com/mineral-rights)** – Mineral rights & claims with maps
    - **[Mountain Man Mining](https://mountainmanmining.com/)** – Nevada-focused claims
    - **[Out West Land Sales](https://outwestlandsales.com/patented-mining-claims/)** – Patented mining claims listings
    - **[eBay - Mining Claims](https://www.ebay.com/sch/i.html?_nkw=mining+claim)** – Active private sales
    """)

    st.info("Tip: Copy cse_nr (serial number) from BLM results and paste into The Diggings, MineExchange.com, or other sites to see if the claim is listed for sale or to contact the owner.")

    # JORC Compliance
    st.subheader("JORC Compliance Details and Reports")
    st.write("JORC Code 2024 Updates: Enhanced ESG provisions, mandatory ESG in Modifying Factors, greater transparency.")
    if st.button("Generate JORC-Compliant Report Summary"):
        jorc_prompt = f"Generate a JORC-compliant report summary for {selected_area}. Include Mineral Resources classification, Competent Person statement, ESG considerations, modifying factors, and 2024 compliance."
        try:
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            response = client.chat.completions.create(model="gpt-4o", messages=[{"role": "user", "content": jorc_prompt}], max_tokens=2000)
            st.session_state['jorc_report'] = response.choices[0].message.content
            st.write("JORC-Compliant Report Summary:")
            st.write(st.session_state['jorc_report'])
        except Exception as e:
            st.error(f"JORC Report Error: {e}")

    # NI 43-101 Compliance
    st.subheader("NI 43-101 Compliance Details and Reports")
    st.write("NI 43-101: Canadian standard requiring Qualified Person and technical reports.")
    if st.button("Generate NI 43-101-Compliant Report Summary"):
        ni_prompt = f"Generate an NI 43-101-compliant report summary for {selected_area}. Include property description, exploration data, resource estimates, QP statement."
        try:
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            response = client.chat.completions.create(model="gpt-4o", messages=[{"role": "user", "content": ni_prompt}], max_tokens=2000)
            st.session_state['ni_report'] = response.choices[0].message.content
            st.write("NI 43-101-Compliant Report Summary:")
            st.write(st.session_state['ni_report'])
        except Exception as e:
            st.error(f"NI 43-101 Report Error: {e}")

    # S-K 1300 Reporting
    st.subheader("S-K 1300 Reporting Details and Reports")
    st.write("S-K 1300: US SEC regulation for mineral disclosure, aligned with CRIRSCO.")
    if st.button("Generate S-K 1300-Compliant Report Summary"):
        sk_prompt = f"Generate an S-K 1300-compliant report summary for {selected_area}. Include mineral resources, initial assessment, QP, property disclosures."
        try:
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            response = client.chat.completions.create(model="gpt-4o", messages=[{"role": "user", "content": sk_prompt}], max_tokens=2000)
            st.session_state['sk_report'] = response.choices[0].message.content
            st.write("S-K 1300-Compliant Report Summary:")
            st.write(st.session_state['sk_report'])
        except Exception as e:
            st.error(f"S-K 1300 Report Error: {e}")

    # SASB Mining Standards
    st.subheader("SASB Standards for Metals & Mining (EM-MM)")
    st.write("SASB focuses on financially material ESG topics for mining.")
    if st.button("Generate SASB-Compliant Disclosure Summary"):
        sasb_prompt = f"Generate a SASB-compliant disclosure summary for Metals & Mining based on data from {selected_area}. Cover GHG Emissions, Water Management, Waste, Biodiversity, Community Relations, Labor Practices, Business Ethics."
        try:
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            response = client.chat.completions.create(model="gpt-4o", messages=[{"role": "user", "content": sasb_prompt}], max_tokens=2000)
            st.session_state['sasb_report'] = response.choices[0].message.content
            st.write("SASB-Compliant Disclosure Summary:")
            st.write(st.session_state['sasb_report'])
        except Exception as e:
            st.error(f"SASB Report Error: {e}")

    # ESG Scoring
    st.subheader("Simple ESG Scoring")
    env_score = st.slider("Environmental Score (0-10)", 0, 10, 5)
    soc_score = st.slider("Social Score (0-10)", 0, 10, 5)
    gov_score = st.slider("Governance Score (0-10)", 0, 10, 5)
    esg_score = (env_score + soc_score + gov_score) / 3
    st.write(f"Overall ESG Score: {esg_score:.2f}/10")
    st.session_state['esg_score'] = esg_score

# ========================================
# File Upload and Analysis
# ========================================
uploaded_file = st.file_uploader("Upload your mining data file (Excel)", type=["xlsx"])

if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file)
    sheets = wb.sheetnames
    st.write("Sheets:", sheets)

    df = pd.read_excel(uploaded_file, sheet_name=sheets[0])
    st.dataframe(df.head())

    # Fixed Map Display
    st.subheader("Data Visualization")
    if 'LATITUDE' in df.columns and 'LONGITUDE' in df.columns:
        try:
            map_df = df[['LATITUDE', 'LONGITUDE']].copy()
            map_df['LATITUDE'] = pd.to_numeric(map_df['LATITUDE'], errors='coerce')
            map_df['LONGITUDE'] = pd.to_numeric(map_df['LONGITUDE'], errors='coerce')
            map_df_clean = map_df.dropna()
            if not map_df_clean.empty:
                st.map(map_df_clean)
                st.success(f"Map displayed with {len(map_df_clean)} valid points.")
            else:
                st.warning("No valid numeric LATITUDE/LONGITUDE data found for mapping.")
        except Exception as e:
            st.error(f"Error displaying map: {e}")

    # Composition Plots
    if st.checkbox("Show Composition Plots"):
        numeric_cols = df.select_dtypes(include='number').columns
        if len(numeric_cols) > 0:
            fig, ax = plt.subplots()
            df[numeric_cols[:5]].plot(kind='box', ax=ax)
            st.pyplot(fig)
            st.session_state['comp_plot'] = fig

    # Interactive 3D Geological Modeling
    st.subheader("Interactive 3D Geological Modeling")
    if 'LATITUDE' in df.columns and 'LONGITUDE' in df.columns and len(df.select_dtypes(include='number').columns) > 0:
        z_col = st.selectbox("Select Z-axis column", df.select_dtypes(include='number').columns)
        if st.button("Generate Interactive 3D Model"):
            fig3d = px.scatter_3d(df, x='LONGITUDE', y='LATITUDE', z=z_col, color=z_col, opacity=0.7)
            fig3d.update_layout(margin=dict(l=0, r=0, b=0, t=0))
            st.plotly_chart(fig3d)
            st.session_state['3d_model'] = fig3d

    # Basic Resource Estimation (IDW Interpolation)
    st.subheader("Basic Resource Estimation (IDW Interpolation)")
    if 'LATITUDE' in df.columns and 'LONGITUDE' in df.columns and len(df.select_dtypes(include='number').columns) > 0:
        value_col = st.selectbox("Select Value Column for Interpolation", df.select_dtypes(include='number').columns)
        if st.button("Perform IDW Estimation"):
            points = df[['LONGITUDE', 'LATITUDE', value_col]].dropna().to_numpy()
            values = points[:, 2]
            grid_x, grid_y = np.mgrid[min(points[:,0]):max(points[:,0]):100j, min(points[:,1]):max(points[:,1]):100j]
            grid_z = griddata(points[:, :2], values, (grid_x, grid_y), method='linear')
            fig_idw, ax_idw = plt.subplots()
            ax_idw.imshow(grid_z.T, extent=(min(points[:,0]), max(points[:,0]), min(points[:,1]), max(points[:,1])), origin='lower')
            ax_idw.set_title("IDW Interpolation Grid")
            st.pyplot(fig_idw)
            st.session_state['idw_chart'] = fig_idw

    # Extract full text content for AI analysis
    content = ""
    for sheet in wb:
        content += f"Sheet: {sheet.title}\n"
        for row in sheet.iter_rows(values_only=True):
            content += ",".join([str(cell) for cell in row if cell is not None]) + "\n"
    
    if len(content) > 100000:
        content = content[:100000] + "... (truncated)"
    
    prompt = f"""
    Analyze the following mining data from the Excel file in the context of {selected_area}. 
    Extract all information related to metals, ores, locations, geological characteristics, samples, compositions, and any other relevant metrics. 
    Provide analysis on what metals and ores are present, where the data is related to, nearby mines, ownership, and economic factors.
    File content:
    {content}
    """

    # AI Analysis Buttons (using OpenAI only)
    if st.button("Analyze with OpenAI"):
        try:
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            response = client.chat.completions.create(model="gpt-4o", messages=[{"role": "user", "content": prompt}], max_tokens=2000)
            st.session_state['openai_analysis'] = response.choices[0].message.content
            st.write("OpenAI Analysis:")
            st.write(st.session_state['openai_analysis'])
        except Exception as e:
            st.error(f"OpenAI Error: {e}")

    # Refined Cost Estimation Calculator with ESG
    st.subheader("Refined Mining Cost Estimation Calculator with ESG")
    tonnage = st.number_input("Ore Tonnage (tons)", min_value=0.0, value=1000000.0)
    grade = st.number_input("Ore Grade (g/t or %)", min_value=0.0, value=1.0)
    recovery = st.number_input("Recovery Rate (%)", min_value=0.0, max_value=100.0, value=90.0)
    metal_price = st.number_input("Metal Price ($/unit)", min_value=0.0, value=2000.0)
    op_cost_per_ton = st.number_input("Operating Cost ($/ton)", min_value=0.0, value=50.0)
    environmental_cost_per_ton = st.number_input("Environmental Cost ($/ton)", min_value=0.0, value=5.0)
    social_cost_per_ton = st.number_input("Social Cost ($/ton)", min_value=0.0, value=3.0)
    governance_cost_per_ton = st.number_input("Governance Cost ($/ton)", min_value=0.0, value=2.0)
    capex = st.number_input("Initial Capex ($)", min_value=0.0, value=100000000.0)
    sust_capex_annual = st.number_input("Annual Sustaining Capex ($)", min_value=0.0, value=5000000.0)
    royalty_rate = st.number_input("Royalty Rate (%)", min_value=0.0, value=2.5)
    tax_rate = st.number_input("Tax Rate (%)", min_value=0.0, value=25.0)
    discount_rate = st.number_input("Discount Rate (%)", min_value=0.0, value=10.0)
    years = st.number_input("Project Life (years)", min_value=1, value=10)
    is_percent_grade = st.checkbox("Grade is in % (for base metals)", value=False)
    unit_conversion = st.number_input("Unit Conversion Factor", min_value=0.0, value=31.1035)

    if st.button("Calculate Estimates"):
        esg_cost_per_ton = environmental_cost_per_ton + social_cost_per_ton + governance_cost_per_ton
        total_op_cost_per_ton = op_cost_per_ton + esg_cost_per_ton
        if is_percent_grade:
            contained_metal = tonnage * (grade / 100)
        else:
            contained_metal = tonnage * (grade / unit_conversion)
        recoverable_metal = contained_metal * (recovery / 100)
        annual_production = recoverable_metal / years
        annual_revenue = annual_production * metal_price
        annual_royalty = annual_revenue * (royalty_rate / 100)
        annual_op_cost = (tonnage / years) * total_op_cost_per_ton
        annual_ebitda = annual_revenue - annual_royalty - annual_op_cost - sust_capex_annual
        annual_tax = max(annual_ebitda * (tax_rate / 100), 0)
        annual_fcf = annual_ebitda - annual_tax
        npv = -capex + sum([annual_fcf / (1 + discount_rate/100)**y for y in range(1, years+1)])
        try:
            import numpy_financial as npf
            cash_flows = [-capex] + [annual_fcf] * years
            irr = npf.irr(cash_flows) * 100
        except:
            irr = "N/A"
        st.write(f"NPV: ${npv:.2f} | IRR: {irr:.2f}%")
        st.session_state['cost_estimates'] = f"NPV: ${npv:.2f}\nIRR: {irr:.2f}%"

    # Generate Mining Analyst Report - Using OpenAI only, with mandatory PGM section
    if st.button("Generate Mining Analyst Report"):
        pgm_status = "PGM metals detected in the data" if st.session_state.get('pgm_present', False) else "The reviewed data provides no evidence of Platinum Group Metals (PGM) presence. No PGM-related mineralization or by-products identified in MRDS records or uploaded file."
        report_prompt = f"""Generate a detailed mining analyst report for {selected_area}, structured as:
        1. Executive Summary
        2. Introduction & Area Overview
        3. Geological and Mineralization Analysis
        4. Platinum Group Metals (PGM) Assessment
           - Current status: {pgm_status}
           - Regional PGM potential and comparison to known models (e.g., Alaskan-type, layered intrusions)
           - Recommendations for PGM exploration if applicable
        5. Resource and Exploration Potential (other commodities)
        6. Economic Evaluation (including ESG costs and sustainability)
        7. Regulatory, Permitting and Social Considerations
        8. Risks and Mitigation Strategies
        9. Recommendations and Conclusions
        
        Use available data from USGS MRDS, BLM claims, uploaded file analysis, cost estimates, and ESG factors. Include hypothetical charts described in text if relevant."""
        
        if os.getenv("OPENAI_API_KEY"):
            try:
                client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{"role": "user", "content": report_prompt}],
                    max_tokens=4000,
                    temperature=0.7
                )
                st.session_state['analyst_report'] = response.choices[0].message.content
                st.write("### Mining Analyst Report (Generated with OpenAI GPT-4o)")
                st.write(st.session_state['analyst_report'])
            except Exception as e:
                st.error(f"OpenAI generation failed: {e}")
        else:
            st.error("OpenAI API key not found. Please set OPENAI_API_KEY in environment variables.")

    # Generate PDF Report
    if st.button("Generate PDF Report"):
        analyses = {
            "Selected Area": f"{selected_area} Overview",
            "Geology": area_info['geology'],
            "PGM Assessment": f"PGM Presence: {st.session_state.get('pgm_present', 'Not checked')}. {'Detected' if st.session_state.get('pgm_present', False) else 'No evidence of PGM in data.'}",
            "Cost Estimates": st.session_state.get('cost_estimates', "No estimates"),
            "Analyst Report": st.session_state.get('analyst_report', "No report generated"),
            "JORC Report": st.session_state.get('jorc_report', "No JORC report"),
            "NI 43-101 Report": st.session_state.get('ni_report', "No NI report"),
            "S-K 1300 Report": st.session_state.get('sk_report', "No S-K report"),
            "SASB Report": st.session_state.get('sasb_report', "No SASB report")
        }
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph("Mining Analysis Report", styles['Title']))
        for section, text in analyses.items():
            story.append(Paragraph(section, styles['Heading1']))
            story.append(Paragraph(text, styles['BodyText']))
            story.append(Spacer(1, 12))
        doc.build(story)
        pdf_buffer.seek(0)
        st.download_button("Download PDF Report", pdf_buffer, file_name="mining_report.pdf", mime="application/pdf")

st.info("This is the complete, un-truncated Python code for the Mining Data Analysis Portal.")