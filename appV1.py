import streamlit as st
import pandas as pd
import openpyxl
from openai import OpenAI
import google.generativeai as genai
import requests
import json
import os
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet
import folium
from streamlit_folium import folium_static
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from PIL import Image as PilImage
import pdfplumber  # Add to install: pip install pdfplumber
import random

# Note: This code assumes you have API keys for OpenAI, Google Gemini, and xAI set as environment variables.
# Install required packages: pip install streamlit pandas openpyxl openai google-generativeai requests reportlab folium streamlit-folium matplotlib pillow pdfplumber
# xAI API is hypothetical; adjust based on actual xAI API documentation (see https://x.ai/api).

st.title("Mining Data Analysis Portal - Enhanced with USGS MRDS & Earth MRI Integration")

# Expanded areas including major US geological regions known for gold, silver, precious metals (PM), and rare earth elements (REE)
mineral_areas = {
    "Northern Rio Grande Rift (Colorado) - Au, Ag, Mo": {
        "description": "Covers areas like Leadville, San Luis Basin, Taos Plateau Volcanic Field. Known for gold, silver, molybdenum.",
        "geology": "Broad downwarp with basins, volcanic features along Jemez Lineament. The rift formed ~36-37 Ma due to crustal extension and thinning. Basins like San Luis are complex, divided by intrabasin horsts, with low-angle faults. Sediments deposited in closed basins under intermittent flooding.",
        "geothermal": "High geothermal potential in linkage zones and basins like San Luis. Heat flow >4.0 HFU in parts; hot springs and geothermal wells indicate resources for power generation. Valles Caldera (central but influential) has high-temperature systems (up to 300°C). Exploration ongoing with potential for EGS (Enhanced Geothermal Systems).",
        "search_query": "mines and volcanic fields in northern Rio Grande Rift Colorado",
        "usgs_query": {"state": "Colorado", "commodity": "gold,silver,molybdenum"},
        "state": "Colorado"
    },
    "Central Rio Grande Rift (New Mexico) - Cu, Pb, Zn, U": {
        "description": "Includes Espanola, Albuquerque, Socorro basins, Jemez Volcanic Field, Cerros del Rio. Rich in copper, lead, zinc, uranium.",
        "geology": "En echelon basins, half-grabens, with mid-Oligocene to Pleistocene volcanism. Española basin: 2-3 km deep, began as downwarp in late Oligocene. Albuquerque-Belen basins with ~0.3 mm/yr extension. Complex basins with horsts; late Oligocene magmatism imprinted thermal boundaries. Natural resources in rift basins.",
        "geothermal": "Significant geothermal resources with volcanics; Valles Caldera and Ojo Caliente hot springs show distal connections. High heat flow suggests vertical fractures for magma/groundwater interaction. Known geothermal areas like Jemez Springs (up to 100°C). Potential for binary cycle plants; assessments indicate moderate-high temperature resources.",
        "search_query": "mines and volcanic fields in central Rio Grande Rift New Mexico",
        "usgs_query": {"state": "New Mexico", "commodity": "copper,lead,zinc,uranium"},
        "state": "New Mexico"
    },
    "Southern Rio Grande Rift (New Mexico/Texas/Mexico) - Au, Ag, Cu": {
        "description": "Potrillo Volcanic Field, Mesilla Basin, extending to Chihuahua. Limited metallic deposits, but cinder and aggregate resources; nearby copper mines like Tyrone.",
        "geology": "Narrow rift segments, monogenetic volcanic fields, Basin and Range extension into Mexico. Rift started ~36 Ma with westerly extension. Basins like Santo Domingo form large accommodation zones. Distributed deformation across rift, Great Plains.",
        "geothermal": "Evaluated in areas like Truth or Consequences with high heat flow anomalies. Self-potential surveys in regions like Radium Springs show potential. Moderate resources with hot springs; under-explored but promising for low-temperature applications. Overall rift anomalies suggest extensive fractures for geothermal fluid circulation.",
        "search_query": "mines and volcanic fields in southern Rio Grande Rift New Mexico Texas Mexico",
        "usgs_query": {"state": "New Mexico", "commodity": "gold,silver,copper"},
        "state": "New Mexico"
    },
    "Carlin Trend (Nevada) - Au": {
        "description": "World-class gold mining district in northern Nevada, known for Carlin-type gold deposits.",
        "geology": "Sedimentary-hosted disseminated gold in Paleozoic rocks, associated with intrusive igneous activity.",
        "geothermal": "Moderate potential due to Basin and Range extension.",
        "search_query": "gold mines in Carlin Trend Nevada",
        "usgs_query": {"state": "Nevada", "commodity": "gold"},
        "state": "Nevada"
    },
    "Black Hills (South Dakota) - Au, Ag": {
        "description": "Historic gold rush area, including Homestake Mine, one of the largest gold producers in US history.",
        "geology": "Precambrian metamorphic rocks with Tertiary intrusions.",
        "geothermal": "Low to moderate.",
        "search_query": "gold silver mines in Black Hills South Dakota",
        "usgs_query": {"state": "South Dakota", "commodity": "gold,silver"},
        "state": "South Dakota"
    },
    "Appalachian Region (Eastern US) - Au, Ag": {
        "description": "Gold and silver in Piedmont and Blue Ridge provinces, e.g., Virginia, North Carolina.",
        "geology": "Metamorphic and volcanic rocks with vein deposits.",
        "geothermal": "Low.",
        "search_query": "gold silver mines in Appalachian US",
        "usgs_query": {"state": "Virginia", "commodity": "gold,silver"},
        "state": "Virginia"
    },
    "Bear Lodge (Wyoming) - REE": {
        "description": "Major rare earth elements deposit in the Black Hills uplift.",
        "geology": "Alkaline igneous complex with carbonatite intrusions.",
        "geothermal": "Low.",
        "search_query": "REE mines in Bear Lodge Wyoming",
        "usgs_query": {"state": "Wyoming", "commodity": "rare earths"},
        "state": "Wyoming"
    },
    "Round Top (Texas) - REE, Li": {
        "description": "Rhyolite-hosted rare earth and lithium deposit.",
        "geology": "Tertiary intrusive rhyolite laccolith.",
        "geothermal": "Moderate.",
        "search_query": "REE lithium mines in Round Top Texas",
        "usgs_query": {"state": "Texas", "commodity": "rare earths,lithium"},
        "state": "Texas"
    },
    "Bokan Mountain (Alaska) - REE, U": {
        "description": "Peralkaline granite-hosted rare earth and uranium.",
        "geology": "Jurassic peralkaline intrusive complex.",
        "geothermal": "High in some Alaskan areas.",
        "search_query": "REE uranium mines in Bokan Mountain Alaska",
        "usgs_query": {"state": "Alaska", "commodity": "rare earths,uranium"},
        "state": "Alaska"
    },
    "Mojave Desert (California) - REE, Au": {
        "description": "Mountain Pass Mine, world's largest REE producer outside China; also gold.",
        "geology": "Carbonatite deposits in Precambrian gneiss.",
        "geothermal": "High in Imperial Valley nearby.",
        "search_query": "REE gold mines in Mojave Desert California",
        "usgs_query": {"state": "California", "commodity": "rare earths,gold"},
        "state": "California"
    },
    "Mother Lode (California) - Au": {
        "description": "Historic California Gold Rush area along Sierra Nevada foothills.",
        "geology": "Mesothermal quartz veins in metamorphic rocks.",
        "geothermal": "Moderate.",
        "search_query": "gold mines in Mother Lode California",
        "usgs_query": {"state": "California", "commodity": "gold"},
        "state": "California"
    },
    "Cripple Creek (Colorado) - Au, Ag": {
        "description": "Volcanic-hosted epithermal gold-silver deposits.",
        "geology": "Oligocene caldera with telluride minerals.",
        "geothermal": "High.",
        "search_query": "gold silver mines in Cripple Creek Colorado",
        "usgs_query": {"state": "Colorado", "commodity": "gold,silver"},
        "state": "Colorado"
    },
    "Comstock Lode (Nevada) - Ag, Au": {
        "description": "Famous silver mining district near Virginia City.",
        "geology": "Epithermal veins in Tertiary volcanics.",
        "geothermal": "High.",
        "search_query": "silver gold mines in Comstock Lode Nevada",
        "usgs_query": {"state": "Nevada", "commodity": "silver,gold"},
        "state": "Nevada"
    },
    "Idaho Batholith (Idaho) - Au, Ag, REE": {
        "description": "Granitic intrusions with vein and placer deposits.",
        "geology": "Cretaceous granites with polymetallic veins.",
        "geothermal": "Moderate.",
        "search_query": "gold silver REE mines in Idaho Batholith",
        "usgs_query": {"state": "Idaho", "commodity": "gold,silver,rare earths"},
        "state": "Idaho"
    },
    "Pebble (Alaska) - Cu, Au, Mo": {
        "description": "Porphyry copper-gold-molybdenum deposit.",
        "geology": "Tertiary intrusive complex.",
        "geothermal": "High.",
        "search_query": "copper gold mines in Pebble Alaska",
        "usgs_query": {"state": "Alaska", "commodity": "copper,gold,molybdenum"},
        "state": "Alaska"
    },
}

# Basin and Range Province info
basin_range_info = {
    "description": "The Basin and Range Province covers a large portion of the southwestern United States and western Mexico, including most of Nevada, parts of California, Oregon, Utah, Arizona, New Mexico, and extending into northern Mexico. It is the most geographically extensive 'young' geologic region in North America.",
    "geology": "Characterized by unique basin and range topography with abrupt elevation changes, alternating narrow faulted mountain chains (ranges) and flat arid valleys (basins). Valleys and mountains run roughly north-south, separated by about 30 km, like continental corduroy. Formed by extensional tectonics, with normal faults pushing up mountains and carving valleys below. Includes the Great Basin with internal drainage (no outflow to sea) and part of the lower Colorado River drainage. Bedrock geology spans broad geologic time and environments. The Rio Grande Rift is a prominent feature within the eastern part of the province."
}

# Colorado Plateau info
colorado_plateau_info = {
    "description": "The Colorado Plateau is a physiographic province covering parts of Arizona, Utah, Colorado, and New Mexico. Known for iconic landmarks like the Grand Canyon, Zion, Arches, and Bryce Canyon National Parks.",
    "geology": "Largely made up of high desert with scattered forests, characterized by flat-lying sedimentary rocks sculpted into mesas, buttes, canyons, and badlands. Stable crustal block, uplifted ~8,500 feet without significant deformation. Composed of Precambrian basement, Paleozoic to Mesozoic sediments (sandstones, limestones, shales), and some igneous/metamorphic rocks. Uplift due to mantle processes; home to ancient landscapes with fossils. Mineral resources include uranium, coal, oil/gas; high potential for potash and rare earths."
}

st.header("Basin and Range Province Overview")
st.write(f"**Description:** {basin_range_info['description']}")
st.write(f"**Geology:** {basin_range_info['geology']}")

st.header("Colorado Plateau Overview")
st.write(f"**Description:** {colorado_plateau_info['description']}")
st.write(f"**Geology:** {colorado_plateau_info['geology']}")

selected_area = st.selectbox("Select Mineral/Geological Area for Analysis", list(mineral_areas.keys()))

if selected_area:
    area_info = mineral_areas[selected_area]
    st.write(f"**Description:** {area_info['description']}")
    st.write(f"**Detailed Geology:** {area_info['geology']}")
    st.write(f"**Geothermal Potential:** {area_info['geothermal']}")
    st.write(f"**Suggested Search Query:** {area_info['search_query']}")

    # Interactive Map Feature
    st.subheader("Interactive Map of Selected Area")
    # Approximate coordinates for areas
    coords = {
        "Northern Rio Grande Rift (Colorado) - Au, Ag, Mo": [37.5, -106.0],
        "Central Rio Grande Rift (New Mexico) - Cu, Pb, Zn, U": [35.0, -106.5],
        "Southern Rio Grande Rift (New Mexico/Texas/Mexico) - Au, Ag, Cu": [32.0, -107.0],
        "Carlin Trend (Nevada) - Au": [40.8, -116.0],
        "Black Hills (South Dakota) - Au, Ag": [44.0, -103.5],
        "Appalachian Region (Eastern US) - Au, Ag": [37.5, -80.0],
        "Bear Lodge (Wyoming) - REE": [44.5, -104.5],
        "Round Top (Texas) - REE, Li": [31.3, -105.5],
        "Bokan Mountain (Alaska) - REE, U": [55.0, -132.0],
        "Mojave Desert (California) - REE, Au": [35.0, -116.0],
        "Mother Lode (California) - Au": [38.5, -120.5],
        "Cripple Creek (Colorado) - Au, Ag": [38.7, -105.2],
        "Comstock Lode (Nevada) - Ag, Au": [39.3, -119.6],
        "Idaho Batholith (Idaho) - Au, Ag, REE": [45.0, -115.0],
        "Pebble (Alaska) - Cu, Au, Mo": [59.7, -155.3],
    }
    area_coord = coords.get(selected_area, [35.0, -106.5])
    m = folium.Map(location=area_coord, zoom_start=7)
    folium.Marker(area_coord, popup=selected_area).add_to(m)
    folium_static(m)

    # USGS Bulletin 1693 Integration
    st.subheader("USGS Bulletin 1693 Integration (Mineral Deposit Models)")
    if st.button("Summarize USGS Bulletin 1693 PDF"):
        try:
            bulletin_url = "https://pubs.usgs.gov/bul/1693/report.pdf"
            response = requests.get(bulletin_url)
            if response.status_code == 200:
                pdf_buffer = io.BytesIO(response.content)
                with pdfplumber.open(pdf_buffer) as pdf:
                    text = ""
                    for page in pdf.pages:
                        text += page.extract_text() or ""
                # Use Gemini to summarize
                genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
                model = genai.GenerativeModel('gemini-1.5-pro-latest')
                summary_prompt = f"Summarize the USGS Bulletin 1693 PDF content: {text[:10000]} Focus on mineral deposit models, especially PGM-related ones. Extract key sections on PGM deposits, characteristics, geological settings, and models for checking PGM presence."
                summary_response = model.generate_content(summary_prompt)
                st.session_state['bulletin_summary'] = summary_response.text
                st.write("USGS Bulletin 1693 Summary:")
                st.write(st.session_state['bulletin_summary'])
            else:
                st.error("Failed to fetch USGS Bulletin 1693 PDF.")
        except Exception as e:
            st.error(f"Bulletin Integration Error: {e}")

    # Expanded USGS MRDS Integration with Deposit Type Analysis
    st.subheader("USGS Mineral Resources Data System (MRDS) Integration")
    if st.button("Query USGS MRDS for Selected Area"):
        try:
            usgs_url = "https://mrdata.usgs.gov/mrds/search.php"
            params = {
                "format": "json",
                "max": 50,  # Increased for fuller results
                **area_info['usgs_query']
            }
            response = requests.get(usgs_url, params=params)
            if response.status_code == 200:
                data = response.json()
                st.session_state['usgs_data'] = data
                # Expanded Parsing
                if 'records' in data and data['records']:
                    usgs_df = pd.DataFrame(data['records'])
                    # Select key columns
                    key_cols = ['mrds_id', 'mas_id', 'site_name', 'latitude', 'longitude', 'region', 'country', 'state', 'county', 'com_type', 'commod1', 'commod2', 'commod3', 'oper_type', 'dep_type', 'prod_size', 'dev_stat', 'ore', 'gangue', 'other_matl', 'orebody_fm', 'work_type', 'model', 'alteration', 'ore_ctrl', 'yrfst_prd', 'yrlst_prd', 'dyfst_dev', 'dylst_dev', 'disc_yr', 'prod_yrs', 'discr']
                    usgs_df = usgs_df[key_cols].dropna(axis=1, how='all')  # Drop empty columns
                    st.dataframe(usgs_df)
                    # Summary
                    st.write("Summary of Commodities:")
                    commodities = usgs_df[['commod1', 'commod2', 'commod3']].melt().value.value_counts().head(10)
                    st.bar_chart(commodities)
                    st.session_state['usgs_df'] = usgs_df
                    st.session_state['usgs_summary_chart'] = commodities
                    # Check for PGM Presence
                    pgm_terms = ['platinum', 'palladium', 'rhodium', 'iridium', 'osmium', 'ruthenium', 'pge', 'pgm']
                    pgm_present = any(any(term in str(commod).lower() for term in pgm_terms) for commod in usgs_df[['commod1', 'commod2', 'commod3']].values.flatten() if commod)
                    st.write(f"PGM Presence Detected: {pgm_present}")
                    if pgm_present:
                        st.write("PGM Commodities Found - Analyzing with Bulletin 1693 Models.")
                        # Expanded PGM Analysis with Detailed Models and Case Studies
                        st.subheader("Detailed PGM Deposit Models and Case Studies")
                        st.write("""
                        Detailed PGM Model Examples (from USGS Bulletin 1693 and similar):
                        1. **Alaskan-type PGE Deposits (Model 9a)**: Associated with zoned mafic-ultramafic intrusions (e.g., dunite, clinopyroxenite). PGE in sulfides like pentlandite, pyrrhotite. Characteristics: High Pd/Pt ratios, formed in arc settings. Case Study: Stillwater Complex (MT) - Layered intrusion with J-M Reef, world's highest-grade PGE deposit, operated by Sibanye-Stillwater, producing ~0.5 Moz PGE annually.
                        2. **Podiform Chromite Deposits (Model 8a)**: PGE as by-product in ophiolites. PGE in laurite inclusions in chromite. Characteristics: Low Pd, high Ru-Ir-Os. Case Study: Josephine Ophiolite (OR-CA) - Small-scale chromite mining with minor PGE recovery; historical production.
                        3. **Stratiform PGE in Layered Intrusions (Model 5b)**: Reef-style in Bushveld-type complexes. PGE in reef horizons with sulfides. Characteristics: High Pt-Pd. Case Study: Duluth Complex (MN) - Potential for PGE in Cu-Ni sulfides; exploration by Antofagasta for Ni-Cu-PGE.
                        4. **Synorogenic-Ni-Cu-PGE (Model 7a)**: In synorogenic intrusions. PGE in massive sulfides. Characteristics: High Ni-Cu with PGE. Case Study: Turnagain (Canada), but US analog in Voisey's Bay-style; limited US examples like Eagle Mine (MI) for Ni-Cu with minor PGE.
                        5. **Flood Basalt-Associated Ni-Cu-PGE (Model 5a)**: In komatiitic flows or sills. Sulfides in reefs. Characteristics: High Pd. Case Study: Columbia River Basalts (US) - Potential in mafic sills; exploration for Ni-Cu-PGE in similar settings.
                        """)
                        # AI PGM Analysis
                        if st.button("AI PGM Deposit Analysis"):
                            pgm_prompt = f"Analyze PGM presence in MRDS data: Commodities: {commodities.to_string()}. Cross-reference with USGS Bulletin 1693 models for PGM deposits (e.g., Alaskan PGE, Podiform Chromite). Provide insights on characteristics, settings, and exploration implications for {selected_area}."
                            try:
                                genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
                                model = genai.GenerativeModel('gemini-1.5-pro-latest')
                                pgm_response = model.generate_content(pgm_prompt)
                                st.session_state['pgm_ai_analysis'] = pgm_response.text
                                st.write("AI PGM Deposit Analysis:")
                                st.write(st.session_state['pgm_ai_analysis'])
                            except Exception as e:
                                st.error(f"AI PGM Analysis Error: {e}")
                    # Check for REE Presence (New)
                    ree_terms = ['lanthanum', 'cerium', 'praseodymium', 'neodymium', 'samarium', 'europium', 'gadolinium', 'terbium', 'dysprosium', 'holmium', 'erbium', 'thulium', 'ytterbium', 'lutetium', 'yttrium', 'scandium', 'ree', 'rare earth']
                    ree_present = any(any(term in str(commod).lower() for term in ree_terms) for commod in usgs_df[['commod1', 'commod2', 'commod3']].values.flatten() if commod)
                    st.write(f"REE Presence Detected: {ree_present}")
                    if ree_present:
                        st.write("REE Commodities Found - Analyzing with Bulletin 1693 Models.")
                        st.subheader("Detailed REE Deposit Models")
                        st.write("""
                        Detailed REE Model Examples (from USGS Bulletin 1693 and similar):
                        1. **Carbonatite Deposits (Model 10)**: REE in apatite, monazite, bastnaesite. Examples: Mountain Pass (CA), Bear Lodge (WY). Characteristics: Alkaline intrusions, high LREE, associated with Nb, U.
                        2. **Peralkaline Granite Deposits (Model 11)**: REE in allanite, zircon, eudialyte. Examples: Bokan Mountain (AK). Characteristics: High HREE, U-Th associated, A-type granites.
                        3. **Phosphorite Deposits (Model 34c)**: REE as by-product in marine phosphates. Examples: Florida phosphorites. Characteristics: Sedimentary, low-grade REE in apatite.
                        4. **Ion-Adsorption Clay Deposits (Model 11d)**: Weathered granites with adsorbed REE. Examples: Chinese deposits, potential in US southeast. Characteristics: Supergene enrichment, easy leaching.
                        5. **Placer Deposits (Model 39a)**: REE in monazite sands. Examples: Idaho placers. Characteristics: Heavy mineral concentrations, beach/river placers.
                        """)
                        # Enhanced REE Visuals
                        if st.checkbox("Show Enhanced REE Visuals"):
                            # Example REE abundance chart
                            ree_elements = ['La', 'Ce', 'Pr', 'Nd', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu', 'Y', 'Sc']
                            example_abundances = [random.uniform(0, 100) for _ in ree_elements]  # Placeholder data
                            fig_ree, ax_ree = plt.subplots()
                            ax_ree.bar(ree_elements, example_abundances)
                            ax_ree.set_title("Example REE Abundance in Deposits")
                            ax_ree.set_ylabel("Abundance (ppm)")
                            st.pyplot(fig_ree)
                            st.session_state['ree_chart'] = fig_ree
                            # Add map or image if available
                            st.image("https://www.usgs.gov/sites/default/files/styles/original/public/2023-02/ree-us-map.png?itok=6yY2z5uK", caption="US REE Deposits Map (USGS)")

                        # AI REE Analysis
                        if st.button("AI REE Deposit Analysis"):
                            ree_prompt = f"Analyze REE presence in MRDS data: Commodities: {commodities.to_string()}. Cross-reference with USGS Bulletin 1693 models for REE deposits (e.g., Carbonatite: alkaline, LREE-rich; Peralkaline Granite: HREE, U-associated; Phosphorite: sedimentary apatite; Ion-Adsorption: weathered clays; Placer: monazite sands). Provide detailed insights on characteristics, settings, examples, and exploration implications for {selected_area}."
                            try:
                                genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
                                model = genai.GenerativeModel('gemini-1.5-pro-latest')
                                ree_response = model.generate_content(ree_prompt)
                                st.session_state['ree_ai_analysis'] = ree_response.text
                                st.write("AI REE Deposit Analysis:")
                                st.write(st.session_state['ree_ai_analysis'])
                            except Exception as e:
                                st.error(f"AI REE Analysis Error: {e}")
                    # MRDS Deposit Type and Model Analysis
                    if 'dep_type' in usgs_df.columns or 'model' in usgs_df.columns:
                        dep_types = usgs_df.get('dep_type', pd.Series()).value_counts()
                        models = usgs_df.get('model', pd.Series()).value_counts()
                        st.write("Deposit Type Analysis:")
                        st.bar_chart(dep_types)
                        st.session_state['deposit_types_chart'] = dep_types
                        st.write("Common Deposit Types: " + ', '.join(dep_types.index[:5]))
                        st.write("Deposit Model Analysis:")
                        st.bar_chart(models)
                        st.session_state['deposit_models_chart'] = models
                        st.write("Common Deposit Models: " + ', '.join(models.index[:5]))
                        # AI Analysis Button
                        if st.button("Analyze MRDS Deposit Models with AI"):
                            ai_prompt = f"Analyze the following MRDS deposit types and models data: Deposit Types: {dep_types.to_string()}\nModels: {models.to_string()}\nProvide insights on common characteristics, economic significance, and relations to geology in {selected_area}."
                            try:
                                genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
                                model = genai.GenerativeModel('gemini-1.5-pro-latest')
                                ai_response = model.generate_content(ai_prompt)
                                st.session_state['mrds_ai_analysis'] = ai_response.text
                                st.write("AI Analysis of MRDS Deposit Models:")
                                st.write(st.session_state['mrds_ai_analysis'])
                            except Exception as e:
                                st.error(f"AI Analysis Error: {e}")
                    else:
                        st.write("No deposit type or model data available in results.")
                    # Map if lat/long present
                    if 'latitude' in usgs_df.columns and 'longitude' in usgs_df.columns:
                        map_df = usgs_df[['latitude', 'longitude', 'site_name']].dropna()
                        if not map_df.empty:
                            mrds_map = folium.Map(location=[map_df['latitude'].mean(), map_df['longitude'].mean()], zoom_start=5)
                            for _, row in map_df.iterrows():
                                folium.Marker([row['latitude'], row['longitude']], popup=row['site_name']).add_to(mrds_map)
                            folium_static(mrds_map)
                else:
                    st.write("No records found.")
            else:
                st.error(f"USGS API Error: {response.status_code}")
        except Exception as e:
            st.error(f"USGS Query Error: {e}")

    # Integrate USGS Earth MRI Data
    st.subheader("USGS Earth Mapping Resources Initiative (Earth MRI) Integration")
    if st.button("Query Earth MRI Data for Selected Area"):
        try:
            # Earth MRI data portal search; no direct API, so use web search proxy or browse
            # For integration, browse the data page and extract summaries
            earth_mri_url = "https://www.usgs.gov/special-topics/earth-mri/data"
            instructions = f"Summarize Earth MRI data releases relevant to {selected_area}, including focus areas, critical minerals, and links to downloads."
            # Hypothetical: Use a tool-like browse, but since no direct, use requests + parse or gemini summarize
            response = requests.get(earth_mri_url)
            if response.status_code == 200:
                # Use Gemini to summarize page content
                genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
                model = genai.GenerativeModel('gemini-1.5-pro-latest')
                summary_prompt = f"Summarize the following page content for Earth MRI data in {selected_area}: {response.text[:10000]}"  # Truncate
                summary_response = model.generate_content(summary_prompt)
                st.session_state['earth_mri_summary'] = summary_response.text
                st.write("Earth MRI Summary:")
                st.write(st.session_state['earth_mri_summary'])
            else:
                st.error("Failed to fetch Earth MRI page.")
        except Exception as e:
            st.error(f"Earth MRI Integration Error: {e}. Note: Earth MRI primarily uses data portals; consider manual download from USGS.")

    # BLM Mining Claims Search
    st.subheader("BLM Mining Claims Search (via The Diggings API)")
    state_code = st.text_input("State Code (e.g., NM for New Mexico, NV for Nevada)", value="NM")
    county = st.text_input("County Name (optional)", value="")
    claim_type = st.selectbox("Claim Type", ["All", "Active", "Closed"])
    if st.button("Search BLM Mining Claims"):
        try:
            # Use The Diggings API for BLM claims (free tier available)
            base_url = "https://thediggings.com/api/search/mining_claims"
            params = {
                "state": state_code.upper(),
                "county": county,
                "status": claim_type.lower() if claim_type != "All" else "",
                "limit": 50
            }
            response = requests.get(base_url, params=params)
            if response.status_code == 200:
                claims_data = response.json()
                if 'results' in claims_data and claims_data['results']:
                    claims_df = pd.DataFrame(claims_data['results'])
                    st.dataframe(claims_df[['claim_id', 'name', 'status', 'type', 'location', 'owner']])
                    st.session_state['blm_claims_df'] = claims_df
                    # Summary
                    st.write("BLM Claims Summary:")
                    status_summary = claims_df['status'].value_counts()
                    st.bar_chart(status_summary)
                    # Map claims if lat/long available
                    if 'latitude' in claims_df.columns and 'longitude' in claims_df.columns:
                        claims_map_df = claims_df[['latitude', 'longitude', 'name']].dropna()
                        if not claims_map_df.empty:
                            blm_map = folium.Map(location=[claims_map_df['latitude'].mean(), claims_map_df['longitude'].mean()], zoom_start=6)
                            for _, row in claims_map_df.iterrows():
                                folium.Marker([row['latitude'], row['longitude']], popup=row['name']).add_to(blm_map)
                            folium_static(blm_map)
                    # AI Analysis for BLM Claims
                    if st.button("Analyze BLM Mining Claims with AI"):
                        ai_prompt = f"Analyze the following BLM mining claims data: {claims_df.to_string()}\nProvide insights on active claims, ownership patterns, and potential for new exploration in {selected_area}."
                        try:
                            genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
                            model = genai.GenerativeModel('gemini-1.5-pro-latest')
                            ai_response = model.generate_content(ai_prompt)
                            st.session_state['blm_ai_analysis'] = ai_response.text
                            st.write("AI Analysis of BLM Mining Claims:")
                            st.write(st.session_state['blm_ai_analysis'])
                        except Exception as e:
                            st.error(f"AI BLM Analysis Error: {e}")
                else:
                    st.write("No BLM claims found for the search criteria.")
            else:
                st.error(f"BLM Claims API Error: {response.status_code}. Note: The Diggings API may require registration for full access.")
        except Exception as e:
            st.error(f"BLM Claims Search Error: {e}. Alternative: Use https://thediggings.com/ or BLM LR2000 system.")

    # BLM MLRS Integration
    st.subheader("BLM MLRS (Mining and Land Records System) Integration")
    if st.button("Query BLM MLRS for Mining Claims"):
        try:
            mlrs_url = "https://mlrs.blm.gov/s/"  # Public viewer
            response = requests.get(mlrs_url)
            if response.status_code == 200:
                # Use Gemini to summarize
                genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
                model = genai.GenerativeModel('gemini-1.5-pro-latest')
                mlrs_prompt = f"Summarize BLM MLRS mining claims data for {area_info.get('state', 'US')} from the page: {response.text[:5000]}. Provide number of active claims, trends, and links to reports or data downloads."
                mlrs_response = model.generate_content(mlrs_prompt)
                st.session_state['mlrs_summary'] = mlrs_response.text
                st.write("BLM MLRS Summary:")
                st.write(st.session_state['mlrs_summary'])
            else:
                st.error("Failed to fetch BLM MLRS page.")
        except Exception as e:
            st.error(f"BLM MLRS Integration Error: {e}. Access via https://mlrs.blm.gov for full data; public API limited.")

    # JORC Compliance Section
    st.subheader("JORC Compliance Details and Reports")
    st.write("JORC Code 2024 Updates: The draft JORC Code was released in August 2024 with enhanced ESG provisions, mandatory ESG considerations in reporting, and updates to Modifying Factors including risk and ESG. The final version is expected in December 2025.")
    if st.button("Generate JORC-Compliant Report Summary"):
        jorc_prompt = f"Generate a JORC-compliant report summary based on the data from {selected_area}. Include sections on Mineral Resources classification (Inferred, Indicated, Measured), Competent Person statement, ESG considerations, modifying factors, and compliance with 2024 updates (transparency, RPEE, risks disclosure)."
        try:
            genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-pro-latest')
            jorc_response = model.generate_content(jorc_prompt)
            st.session_state['jorc_report'] = jorc_response.text
            st.write("JORC-Compliant Report Summary:")
            st.write(st.session_state['jorc_report'])
        except Exception as e:
            st.error(f"JORC Report Generation Error: {e}")

    # NI 43-101 Compliance Section
    st.subheader("NI 43-101 Compliance Details and Reports")
    st.write("NI 43-101 Standards: Latest consolidation June 9, 2023. Proposed repeal and replacement with comment period closing October 10, 2025, including removing technical report requirement for royalty-only issuers.")
    if st.button("Generate NI 43-101-Compliant Report Summary"):
        ni_prompt = f"Generate an NI 43-101-compliant report summary for {selected_area}. Include sections on Mineral Property Description, Exploration Data, Mineral Resource Estimates, Qualified Person statement, and compliance with latest standards."
        try:
            genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-pro-latest')
            ni_response = model.generate_content(ni_prompt)
            st.session_state['ni_report'] = ni_response.text
            st.write("NI 43-101-Compliant Report Summary:")
            st.write(st.session_state['ni_report'])
        except Exception as e:
            st.error(f"NI 43-101 Report Generation Error: {e}")

    # S-K 1300 Reporting Section
    st.subheader("S-K 1300 Reporting Details and Reports")
    st.write("S-K 1300: US SEC regulation for mineral resource reporting, effective since 2019. Requires disclosure of mineral resources based on qualified person's initial assessment, allows inclusion of exploration results. Compliance involves detailed technical reports, Qualified Person oversight, and alignment with CRIRSCO standards.")
    if st.button("Generate S-K 1300-Compliant Report Summary"):
        sk_prompt = f"Generate an S-K 1300-compliant report summary for {selected_area}. Include sections on Mineral Resources, Initial Assessment, Qualified Person, and compliance with S-K 1300 requirements (e.g., property disclosures, resource classification)."
        try:
            genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-pro-latest')
            sk_response = model.generate_content(sk_prompt)
            st.session_state['sk_report'] = sk_response.text
            st.write("S-K 1300-Compliant Report Summary:")
            st.write(st.session_state['sk_report'])
        except Exception as e:
            st.error(f"S-K 1300 Report Generation Error: {e}")

    # SASB Mining Standards Section
    st.subheader("SASB Standards for Metals & Mining (EM-MM)")
    st.write("SASB Standards (under ISSB, latest 2023 with proposed amendments July 2025) for Metals & Mining focus on financially material ESG topics. Key disclosure topics include:")
    st.write("""
    - **Greenhouse Gas Emissions**: Scope 1 and 2 emissions, energy management.
    - **Air Quality**: Emissions of NOx, SOx, particulate matter, mercury.
    - **Energy Management**: Energy consumption, renewable energy use.
    - **Water Management**: Water withdrawal, consumption, recycling in water-stressed areas.
    - **Waste & Hazardous Materials Management**: Tailings storage facilities management, waste generated.
    - **Biodiversity Impacts**: Operations in or near protected areas, impacts on biodiversity.
    - **Community Relations**: Rights of indigenous peoples, community engagement.
    - **Human Rights & Rights of Indigenous Peoples**: Due diligence, conflict-affected areas.
    - **Labor Practices**: Workforce health & safety (fatalities, near misses), labor relations.
    - **Business Ethics & Transparency**: Anti-corruption, payments to governments.
    - **Security, Human Rights & Rights of Indigenous Peoples**: Security practices aligned with Voluntary Principles.
    """)
    if st.button("Generate SASB-Compliant Disclosure Summary"):
        sasb_prompt = f"Generate a SASB-compliant disclosure summary for Metals & Mining based on data from {selected_area}. Cover key topics: GHG Emissions, Water Management, Waste (incl. tailings), Biodiversity, Community Relations, Labor Practices, and Business Ethics. Use available data and estimates."
        try:
            genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-pro-latest')
            sasb_response = model.generate_content(sasb_prompt)
            st.session_state['sasb_report'] = sasb_response.text
            st.write("SASB-Compliant Disclosure Summary:")
            st.write(st.session_state['sasb_report'])
        except Exception as e:
            st.error(f"SASB Report Generation Error: {e}")

uploaded_file = st.file_uploader("Upload your mining data file (Excel)", type=["xlsx"])

if uploaded_file is not None:
    # Read the file using openpyxl
    wb = openpyxl.load_workbook(uploaded_file)
    sheets = wb.sheetnames
    
    st.write("Sheets in the file:", sheets)
    
    # Read the first sheet as dataframe for display and visualization
    df = pd.read_excel(uploaded_file, sheet_name=sheets[0])
    
    st.dataframe(df.head())
    
    # Data Visualization Feature
    st.subheader("Data Visualization")
    if 'LATITUDE' in df.columns and 'LONGITUDE' in df.columns:
        st.map(df[['LATITUDE', 'LONGITUDE']].dropna())
    if st.checkbox("Show Composition Plots"):
        numeric_cols = df.select_dtypes(include='number').columns
        if len(numeric_cols) > 0:
            fig, ax = plt.subplots()
            df[numeric_cols[:5]].plot(kind='box', ax=ax)  # Example plot
            st.pyplot(fig)
            st.session_state['comp_plot'] = fig

    # Interactive 3D Geological Modeling with Plotly
    st.subheader("Interactive 3D Geological Modeling")
    if 'LATITUDE' in df.columns and 'LONGITUDE' in df.columns and len(numeric_cols) > 0:
        z_col = st.selectbox("Select Z-axis column", numeric_cols)
        if st.button("Generate Interactive 3D Model"):
            fig3d = px.scatter_3d(
                df,
                x='LONGITUDE',
                y='LATITUDE',
                z=z_col,
                color=z_col,
                size_max=18,
                opacity=0.7
            )
            fig3d.update_layout(margin=dict(l=0, r=0, b=0, t=0))
            st.plotly_chart(fig3d)
            st.session_state['3d_model'] = fig3d  # Save for report if needed

    # Resource Estimation with IDW
    st.subheader("Basic Resource Estimation (IDW Interpolation)")
    if 'LATITUDE' in df.columns and 'LONGITUDE' in df.columns and len(numeric_cols) > 0:
        value_col = st.selectbox("Select Value Column for Interpolation", numeric_cols)
        if st.button("Perform IDW Estimation"):
            points = df[['LONGITUDE', 'LATITUDE', value_col]].dropna().to_numpy()
            values = points[:, 2]
            grid_x, grid_y = np.mgrid[
                min(points[:,0]):max(points[:,0]):100j,
                min(points[:,1]):max(points[:,1]):100j
            ]
            grid_z = griddata(points[:, :2], values, (grid_x, grid_y), method='linear')  # Approx IDW
            fig_idw, ax_idw = plt.subplots()
            ax_idw.imshow(grid_z.T, extent=(min(points[:,0]), max(points[:,0]), min(points[:,1]), max(points[:,1])), origin='lower')
            ax_idw.set_title("IDW Interpolation Grid")
            st.pyplot(fig_idw)
            st.session_state['idw_chart'] = fig_idw

    # ESG Scoring
    st.subheader("Simple ESG Scoring")
    env_score = st.slider("Environmental Score (0-10)", 0, 10, 5)
    soc_score = st.slider("Social Score (0-10)", 0, 10, 5)
    gov_score = st.slider("Governance Score (0-10)", 0, 10, 5)
    esg_score = (env_score + soc_score + gov_score) / 3
    st.write(f"Overall ESG Score: {esg_score:.2f}/10")
    st.session_state['esg_score'] = esg_score

    # Extract full text content for AI analysis
    content = ""
    for sheet in wb:
        content += f"Sheet: {sheet.title}\n"
        for row in sheet.iter_rows(values_only=True):
            content += ",".join([str(cell) for cell in row if cell is not None]) + "\n"
    
    # Truncate if too long to avoid token limits
    if len(content) > 100000:
        content = content[:100000] + "... (truncated)"
    
    # Incorporate selected area into prompt, with additional mining decision factors
    prompt = f"""
    Analyze the following mining data from the Excel file in the context of {selected_area} along the Rio Grande Rift, including Basin and Range Province and Colorado Plateau influences. 
    Extract all information related to metals, ores, locations, geological characteristics, samples, compositions, and any other relevant metrics. 
    Provide analysis on what metals and ores are present (e.g., Platinum Group Metals, copper, zinc, etc.), where the data is related to (e.g., Las Cruces, New Mexico), 
    if there are nearby mines, who owns them, and if any are for sale. Include economic factors like market prices, mining costs, regulations, environmental impacts, permit processes, and risk assessments.
    Use your knowledge and search capabilities if needed for additional context like nearby mines.
    Consider extensions into adjacent states or regions based on geological formations.

    File content:
    {content}
    """
    
    # OpenAI Analysis Section
    if st.button("Analyze with OpenAI"):
        try:
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            response = client.chat.completions.create(
                model="gpt-4o",  # Or latest model
                messages=[{"role": "user", "content": prompt}],
                max_tokens=2000
            )
            st.session_state['openai_analysis'] = response.choices[0].message.content
            st.write("OpenAI Analysis:")
            st.write(st.session_state['openai_analysis'])
        except Exception as e:
            st.error(f"OpenAI Error: {e}. Ensure API key is set.")
    
    # Google Gemini Analysis Section
    if st.button("Analyze with Google Gemini"):
        try:
            genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-pro-latest')
            response = model.generate_content(prompt)
            st.session_state['gemini_analysis'] = response.text
            st.write("Google Gemini Analysis:")
            st.write(st.session_state['gemini_analysis'])
        except Exception as e:
            st.error(f"Gemini Error: {e}. Ensure API key is set.")
    
    # xAI Grok Analysis Section (Hypothetical API - adjust per xAI docs)
    if st.button("Analyze with xAI Grok"):
        try:
            api_key = os.getenv("XAI_API_KEY")
            url = "https://api.x.ai/v1/chat/completions"  # Hypothetical; check https://x.ai/api for actual
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            data = {
                "model": "grok-4",  # Or appropriate model
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 2000
            }
            response = requests.post(url, headers=headers, json=data)
            if response.status_code == 200:
                result = response.json()
                st.session_state['grok_analysis'] = result['choices'][0]['message']['content']
                st.write("xAI Grok Analysis:")
                st.write(st.session_state['grok_analysis'])
            else:
                st.error(f"xAI API Error: {response.status_code} - {response.text}")
        except Exception as e:
            st.error(f"xAI Error: {e}. Ensure API key is set and endpoint is correct.")

    # Additional: Search for Nearby Mines (Using Gemini as proxy for search)
    if st.button("Search Nearby Mines and Ownership in Selected Area"):
        search_prompt = f"Provide details on mines in {selected_area}, including metals/ores, owners, if for sale, production metrics, reserves, economic viability, regulations, environmental impacts."
        try:
            genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-pro-latest')
            response = model.generate_content(search_prompt)
            st.session_state['mines_info'] = response.text
            st.write("Nearby Mines Information:")
            st.write(st.session_state['mines_info'])
        except Exception as e:
            st.error(f"Search Error: {e}")

    # Refined Cost Estimation Calculator with ESG Costs
    st.subheader("Refined Mining Cost Estimation Calculator with ESG")
    tonnage = st.number_input("Ore Tonnage (tons)", min_value=0.0, value=1000000.0)
    grade = st.number_input("Ore Grade (g/t or %)", min_value=0.0, value=1.0)
    recovery = st.number_input("Recovery Rate (%)", min_value=0.0, max_value=100.0, value=90.0)
    metal_price = st.number_input("Metal Price ($/unit)", min_value=0.0, value=2000.0)
    op_cost_per_ton = st.number_input("Operating Cost ($/ton)", min_value=0.0, value=50.0)
    environmental_cost_per_ton = st.number_input("Environmental Cost ($/ton) (e.g., 5-15% of op)", min_value=0.0, value=5.0)  # Avg from research
    social_cost_per_ton = st.number_input("Social Cost ($/ton) (e.g., 5-10%)", min_value=0.0, value=3.0)
    governance_cost_per_ton = st.number_input("Governance Cost ($/ton) (e.g., 2-5%)", min_value=0.0, value=2.0)
    capex = st.number_input("Initial Capex ($)", min_value=0.0, value=100000000.0)
    sust_capex_annual = st.number_input("Annual Sustaining Capex ($)", min_value=0.0, value=5000000.0)
    royalty_rate = st.number_input("Royalty Rate (%)", min_value=0.0, value=2.5)
    tax_rate = st.number_input("Tax Rate (%)", min_value=0.0, value=25.0)
    discount_rate = st.number_input("Discount Rate (%)", min_value=0.0, value=10.0)
    years = st.number_input("Project Life (years)", min_value=1, value=10)
    is_percent_grade = st.checkbox("Grade is in % (for base metals)", value=False)
    unit_conversion = st.number_input("Unit Conversion Factor (e.g., 31.1035 g/oz)", min_value=0.0, value=31.1035)

    if st.button("Calculate Estimates"):
        esg_cost_per_ton = environmental_cost_per_ton + social_cost_per_ton + governance_cost_per_ton
        total_op_cost_per_ton = op_cost_per_ton + esg_cost_per_ton
        if is_percent_grade:
            contained_metal = tonnage * (grade / 100)
        else:
            contained_metal = tonnage * (grade / unit_conversion)
        recoverable_metal = contained_metal * (recovery / 100)
        annual_production = recoverable_metal / years
        annual_revenue = annual_production * metal_price
        annual_royalty = annual_revenue * (royalty_rate / 100)
        annual_op_cost = (tonnage / years) * total_op_cost_per_ton
        annual_ebitda = annual_revenue - annual_royalty - annual_op_cost - sust_capex_annual
        annual_tax = max(annual_ebitda * (tax_rate / 100), 0)
        annual_fcf = annual_ebitda - annual_tax
        npv = -capex + sum([annual_fcf / (1 + discount_rate/100)**y for y in range(1, years+1)])
        try:
            import numpy_financial as npf
            cash_flows = [-capex] + [annual_fcf] * years
            irr = npf.irr(cash_flows) * 100
        except:
            irr = "N/A"
        st.write(f"Contained Metal: {contained_metal:.2f} units")
        st.write(f"Recoverable Metal: {recoverable_metal:.2f} units")
        st.write(f"Annual Revenue: ${annual_revenue:.2f}")
        st.write(f"Annual Op Cost (incl ESG): ${annual_op_cost:.2f}")
        st.write(f"Annual EBITDA: ${annual_ebitda:.2f}")
        st.write(f"Annual FCF (post-tax): ${annual_fcf:.2f}")
        st.write(f"NPV: ${npv:.2f}")
        st.write(f"IRR: {irr:.2f}%")
        st.session_state['cost_estimates'] = f"Contained: {contained_metal:.2f}\nRecoverable: {recoverable_metal:.2f}\nAnnual Rev: ${annual_revenue:.2f}\nAnnual Op Cost (incl ESG): ${annual_op_cost:.2f}\nEBITDA: ${annual_ebitda:.2f}\nFCF: ${annual_fcf:.2f}\nNPV: ${npv:.2f}\nIRR: {irr:.2f}%"

    # Generate Mining Analyst Report
    if st.button("Generate Mining Analyst Report"):
        report_prompt = f"Generate a detailed mining analyst report for {selected_area}, structured as: Executive Summary, Introduction, Geological Analysis, Economic Evaluation (including ESG costs), Risks and Mitigation (with ESG focus), Recommendations. Incorporate MRDS, BLM, cost estimates, ESG factors, and competitor analysis."
        try:
            genai.configure(api_key=os.getenv("GOOGLE_GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-pro-latest')
            report_response = model.generate_content(report_prompt)
            st.session_state['analyst_report'] = report_response.text
            st.write("Mining Analyst Report:")
            st.write(st.session_state['analyst_report'])
        except Exception as e:
            st.error(f"Report Generation Error: {e}")

# To run: streamlit run this_file.py
# Note: For production, handle large files, add error checking, and secure API keys.